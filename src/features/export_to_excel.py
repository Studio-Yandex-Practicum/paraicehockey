import datetime
import logging
import os

import openpyxl
import requests

from src.core.prometheus_constants import (METRIC_GROUPS, METRIC_NAMES,
                                           PROMETHEUS_URL)
from src.core.settings import settings
from src.core.utils import is_admin

logger = logging.getLogger(__name__)


def get_metrics_list():
    """Получаем группы и их метрики."""
    metrics = {}
    for group in METRIC_GROUPS:
        url = f'{PROMETHEUS_URL}/api/v1/series?match[]={{group="{group}"}}'
        try:
            response = requests.get(url)
        except requests.exceptions.RequestException:
            logging.error('Prometheus: не удалось получить метрику. ')
        data = response.json()['data']
        for metric in data:
            name = metric['__name__']
            if metric['job'] != settings.bot_prometheus:
                continue
            if '_created' in name:
                continue
            if group not in metrics:
                metrics[group] = [name]
            else:
                metrics[group].append(name)
    return metrics


def create_excel_file(start, end, filename, metrics, interval):
    """Создаем excel файл."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = (
        f'Статистика с '
        f'{datetime.datetime.fromtimestamp(start).strftime("%d/%m/%Y %H:%M")} '
        f'по {datetime.datetime.fromtimestamp(end).strftime("%d/%m/%Y %H:%M")}'
    )
    ws.column_dimensions['A'].width = 50

    row = 2  # noqa
    for group in metrics:
        ws.cell(row=row, column=1, value=METRIC_GROUPS[group])
        ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
        row += 1
        for metric in metrics[group]:
            if metric in METRIC_NAMES:
                name = METRIC_NAMES.get(metric, metric)
            else:
                name = metric
            if interval == 'all':
                query = (
                    f'{PROMETHEUS_URL}/api/v1/query?query={metric}'
                    f'{{job="{settings.bot_prometheus}"}}'
                    f'&time={end}&start={start}'
                )
            else:
                query = (
                    f'{PROMETHEUS_URL}/api/v1/query?query=changes({metric}'
                    f'{{job="{settings.bot_prometheus}"}}'
                    f'[{interval}])&time={end}&start={start}'
                )
            try:
                response = requests.get(query)
            except requests.exceptions.RequestException:
                logging.error('Prometheus: не удалось получить метрику.')
            data = response.json()['data']['result']
            for result in data:
                value = result['value'][1]
                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=1)
                ws.cell(row=row, column=2, value=value)
                row += 1
    wb.save(filename)


def export_statistics(start, end, chat_id, context, interval):
    """Экспортируем статистику и сохраняем в телеграмм."""
    filename = (
        f'export_{datetime.datetime.fromtimestamp(start).strftime("%Y-%m-%d")}'
        '.xlsx'
    )
    metrics = get_metrics_list()
    create_excel_file(start, end, filename, metrics, interval)
    bot = context.bot
    with open(filename, 'rb') as f:
        bot.send_document(chat_id=chat_id, document=f)
    os.remove(filename)
    logger.info('Файл успешно сохранен: %s', filename)


@is_admin
def export_for_day(update, context):
    """Экспортируем статистику за день."""
    chat_id = update.message.chat_id
    start = int(
        (datetime.datetime.now() - datetime.timedelta(days=1)).timestamp()
    )
    end = int(datetime.datetime.now().timestamp())
    export_statistics(start, end, chat_id, context, '24h')
    logger.info('Экспорт статистики за день')


@is_admin
def export_for_week(update, context):
    """Экспортируем статистику за 2 недели."""
    chat_id = update.message.chat_id
    start = int(
        (datetime.datetime.now() - datetime.timedelta(weeks=2)).timestamp()
    )
    end = int(datetime.datetime.now().timestamp())
    export_statistics(start, end, chat_id, context, '2w')
    logger.info('Экспорт статистики за неделю')


@is_admin
def export_for_all(update, context):
    """Экспортируем всю статистику."""
    chat_id = update.message.chat_id
    start = int(
        (datetime.datetime.now() - datetime.timedelta(weeks=360)).timestamp()
    )
    end = int(datetime.datetime.now().timestamp())
    export_statistics(start, end, chat_id, context, 'all')
    logger.info('Экспорт всей статистики')
